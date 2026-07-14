from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment

root = Path(__file__).resolve().parents[1]
out_path = root / 'google-apps-script' / 'aff-partners-info' / 'anonymized_table.xlsx'

wb = Workbook()
ws = wb.active
ws.title = 'Synthetic Summary'
ws['A1'] = 'unique_key'
ws['B1'] = 'partner_name'
ws['C1'] = 'campaign_id'
ws['D1'] = 'status'

for row in [
    ['alpha-001', 'Example Partner', 'CMP-100', 'Approved'],
    ['alpha-002', 'Sample Partner', 'CMP-101', 'Pending'],
]:
    ws.append(row)

ws['A2'] = 'alpha-001'
ws['B2'] = 'Example Partner'
ws['C2'] = 'CMP-100'
ws['D2'] = 'Approved'
ws['A3'] = 'alpha-002'
ws['B3'] = 'Sample Partner'
ws['C3'] = 'CMP-101'
ws['D3'] = 'Pending'

ws.cell(row=2, column=5, value='note').comment = Comment('Synthetic data only', 'Automation Bot')
ws.freeze_panes = 'A2'
ws.print_title_rows = '1:1'

ws2 = wb.create_sheet('Hidden Sheet')
ws2['A1'] = 'hidden'
ws2.sheet_state = 'hidden'

wb.create_named_range('synthetic_range', ws, 'A1:D3')
wb.properties.creator = 'Automation Bot'
wb.properties.lastModifiedBy = 'Automation Bot'
wb.properties.title = 'Synthetic Portfolio Workbook'
wb.properties.subject = 'Portfolio Example'
wb.properties.description = 'Synthetic workbook for portfolio demonstration'

wb.save(out_path)
print(out_path)
