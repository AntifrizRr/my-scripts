import pathlib
import re
import subprocess
import sys
from typing import Iterable, List, Tuple

ROOT = pathlib.Path(__file__).resolve().parents[1]
ALLOWED_SUFFIXES = {'.md', '.txt', '.json', '.gs', '.js', '.py', '.yml', '.yaml'}
SKIP_DIRS = {'.git', '.venv', '__pycache__', '.pytest_cache', '.ruff_cache', '.mypy_cache', '.idea'}
SKIP_PATH_PARTS = {'node_modules', 'dist', 'build', '.eggs', 'egg-info', 'logs', '.cache'}
SKIP_FILES = {'tools/secret_scan_test_data.py'}
SAFE_PLACEHOLDERS = {
    'xoxb-example',
    'xoxp-example',
    'xoxa-example',
    'xoxr-example',
    'example.invalid',
    'REPLACE_WITH_SPREADSHEET_ID',
    'example-spreadsheet-id',
    'example',
    'synthetic-secret',
    'synthetic-secret-value',
    '<@example>',
    '<@ops>',
    '<@analyst>',
    '<@synthetic-user>',
}

PASSWORD_ASSIGNMENT_RE = re.compile(r'(?i)\b(?P<name>password|passwd|pwd)\b(?P<type>\s*:\s*[^=\n]+)?\s*=\s*(?P<value>.+)')
HIGH_CONFIDENCE_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    ("slack-webhook", re.compile(r'https://hooks\.slack\.com/[A-Za-z0-9/_-]+')),
    ("slack-user-id", re.compile(r'<@U[A-Z0-9]+>')),
    ("xox-token", re.compile(r'xox[baprs]-[A-Za-z0-9-]+')),
    ("totp-secret", re.compile(r'(?i)totp\s*[:=]\s*[^\s\"]+')),
    ("real-email", re.compile(r'(?i)\b[A-Z0-9._%+-]+@(?:gmail|outlook|mail|yandex|protonmail)\.com\b')),
]

SAFE_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    ("placeholder-spreadsheet-id", re.compile(r'REPLACE_WITH_SPREADSHEET_ID|example-spreadsheet-id')),
    ("placeholder-token", re.compile(r'xox[baprs]-example')),
    ("placeholder-slack-id", re.compile(r'<@(?:example|ops|analyst|synthetic-user)>')),
    ("placeholder-webhook", re.compile(r'https://example\.invalid/webhook')),
    ("placeholder-email", re.compile(r'(?i)example@(?:gmail|outlook|mail|yandex|protonmail)\.com')),
]


def iter_git_files(root: pathlib.Path) -> Iterable[pathlib.Path]:
    proc = subprocess.run(
        ["git", "ls-files", "--cached", "--others", "--exclude-standard"],
        cwd=root,
        capture_output=True,
        text=True,
        check=True,
    )
    for rel in proc.stdout.splitlines():
        rel = rel.strip()
        if not rel:
            continue

        rel_posix = pathlib.Path(rel).as_posix()
        if rel_posix in SKIP_FILES:
            continue

        path = (root / rel).resolve()
        if not path.exists() or not path.is_file():
            continue
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        if any(part in SKIP_PATH_PARTS for part in path.parts):
            continue
        if path.is_symlink():
            continue
        if path.suffix.lower() in {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.pdf', '.zip', '.gz', '.tar', '.whl', '.pyd', '.dll', '.so'}:
            continue
        suffix = path.suffix.lower()
        if suffix and suffix not in ALLOWED_SUFFIXES:
            continue
        yield path


def is_safe_placeholder(value: str) -> bool:
    value = value.strip()
    if not value:
        return False
    if value in SAFE_PLACEHOLDERS:
        return True
    for _, pattern in SAFE_PATTERNS:
        if pattern.search(value):
            return True
    return False


def looks_like_embedded_value(value: str) -> bool:
    value = value.strip()
    if not value:
        return False
    if value in {'password', 'passwd', 'pwd'}:
        return False
    if re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', value):
        return False
    if value.startswith(('"', "'")) and value.endswith(('"', "'")):
        return True
    if value.startswith(('f"', "f'", 'r"', "r'")) and value.endswith(('"', "'")):
        return True
    if re.search(r'["\']', value):
        return True
    return any(token in value for token in ('://', 'xox', '@', 'http'))


def scan_text_file(path: pathlib.Path) -> List[Tuple[int, str, str]]:
    try:
        text = path.read_text(encoding='utf-8', errors='ignore')
    except Exception:
        return []
    findings: List[Tuple[int, str, str]] = []
    for line_no, line in enumerate(text.splitlines(), start=1):
        if 're.compile(' in line or 'patterns =' in line or 'SAFE_PATTERNS' in line:
            continue
        for name, pattern in HIGH_CONFIDENCE_PATTERNS:
            for match in pattern.finditer(line):
                value = match.group(0)
                if is_safe_placeholder(value):
                    continue
                findings.append((line_no, name, value))
        password_match = PASSWORD_ASSIGNMENT_RE.search(line)
        if password_match and looks_like_embedded_value(password_match.group('value')):
            value = password_match.group('value').strip()
            if not is_safe_placeholder(value):
                findings.append((line_no, 'password-assignment', value))
    return findings


def scan_workbook_metadata(root: pathlib.Path) -> List[Tuple[str, str]]:
    path = root / 'google-apps-script' / 'aff-partners-info' / 'anonymized_table.xlsx'
    if not path.exists():
        return []
    try:
        import openpyxl
    except Exception:
        return []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    findings: List[Tuple[str, str]] = []
    if workbook.sheetnames and workbook.sheetnames != ['Synthetic Summary']:
        findings.append(('sheetnames', ','.join(workbook.sheetnames)))
    if workbook.properties.creator and 'example' not in str(workbook.properties.creator).lower():
        findings.append(('creator', str(workbook.properties.creator)))
    if workbook.properties.lastModifiedBy and 'example' not in str(workbook.properties.lastModifiedBy).lower():
        findings.append(('lastModifiedBy', str(workbook.properties.lastModifiedBy)))
    return findings


def main() -> int:
    findings: List[Tuple[pathlib.Path, int, str, str]] = []
    for path in iter_git_files(ROOT):
        for line_no, name, value in scan_text_file(path):
            rel = path.relative_to(ROOT)
            findings.append((rel, line_no, name, value))
    for name, value in scan_workbook_metadata(ROOT):
        findings.append((pathlib.Path('google-apps-script/aff-partners-info/anonymized_table.xlsx'), 0, name, value))

    if findings:
        for rel, line_no, name, value in sorted(findings):
            if line_no:
                print(f'{rel}:{line_no}:{name}:{value}')
            else:
                print(f'{rel}:0:{name}:{value}')
        return 1
    print('No high-confidence secret findings found.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
